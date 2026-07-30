"""
Etapa 8: matriz de criterios de exito CE-01 a CE-20.

Este archivo recorre los veinte criterios de la seccion 5 de la planeacion en el
mismo orden y con la misma numeracion, de modo que la salida de `pytest -v` puede
leerse como evidencia directa: cada nombre de prueba dice a que criterio
corresponde.

No sustituye a las pruebas por etapa (que cubren casos limite en detalle); es la
comprobacion de extremo a extremo de lo que el MVP prometio.
"""

import re

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

import dependencias
import main
from db import crear_tablas_para_pruebas


@pytest.fixture(autouse=True)
def secreto_de_prueba(monkeypatch):
    monkeypatch.setenv("SESSION_SECRET", "secreto_solo_para_pruebas_1234567890")


@pytest.fixture()
def cliente() -> TestClient:
    motor = crear_tablas_para_pruebas()

    def sesion_de_prueba():
        sesion = Session(motor)
        try:
            yield sesion
        finally:
            sesion.close()

    main.app.dependency_overrides[dependencias.obtener_sesion] = sesion_de_prueba
    with TestClient(main.app) as cliente:
        yield cliente
    main.app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# Ayudas del escenario KOVA
# ---------------------------------------------------------------------------

TALLAS = "XXS, XS, S, M, L, XL, XXL"
COLORES = "negro, blanco, azul, verde, rojo, gris, beige"


def crear_cuenta(cliente, empresa="KOVA", correo="dueno@kova.com"):
    return cliente.post(
        "/registro",
        data={"empresa": empresa, "correo": correo, "password": "clave_seguraKOVA"},
        follow_redirects=True,
    )


def crear_producto(cliente, **cambios):
    datos = {
        "nombre": "Polo Premium",
        "categoria": "Playeras",
        "unidad": "pieza",
        "costo": 250,
        "minimo": 10,
        "atributo_1": "Talla",
        "valores_1": TALLAS,
        "atributo_2": "Color",
        "valores_2": COLORES,
    }
    datos.update(cambios)
    return cliente.post("/productos/nuevo", data=datos, follow_redirects=True)


def ids_de_variante(html: str) -> list[str]:
    return re.findall(r'name="cantidad_([0-9a-f-]+)"', html)


def id_de_producto(html: str) -> str:
    return re.search(r'/productos/([0-9a-f-]+)/existencias', html).group(1)


def cargar(cliente, id_producto, cantidades: dict[str, int]):
    datos = {f"cantidad_{v}": str(c) for v, c in cantidades.items()}
    datos["motivo"] = "Inventario inicial"
    return cliente.post(
        f"/productos/{id_producto}/existencias", data=datos, follow_redirects=True
    )


def escenario_kova(cliente) -> tuple[str, list[str]]:
    """Cuenta + polo de 49 variantes + existencias en las primeras tres."""
    crear_cuenta(cliente)
    pantalla = crear_producto(cliente)
    id_producto = id_de_producto(pantalla.text)
    variantes = ids_de_variante(pantalla.text)
    cargar(cliente, id_producto, {variantes[0]: 20, variantes[1]: 5, variantes[2]: 8})
    return id_producto, variantes


# ---------------------------------------------------------------------------
# CE-01 a CE-05
# ---------------------------------------------------------------------------


def test_ce01_registro_multiempresa_y_aislamiento(cliente):
    crear_cuenta(cliente)
    crear_producto(cliente, nombre="Polo de KOVA")
    cliente.post("/salir")

    crear_cuenta(cliente, empresa="Panaderia", correo="dueno@pan.com")
    inventario = cliente.get("/inventario").text

    assert "Polo de KOVA" not in inventario
    assert "Agregar primer producto" in inventario


def test_ce02_onboarding_producto_por_producto(cliente):
    crear_cuenta(cliente)
    assert "Agregar primer producto" in cliente.get("/inventario").text

    pantalla = crear_producto(cliente)
    assert "Captura las existencias" in pantalla.text

    id_producto = id_de_producto(pantalla.text)
    variantes = ids_de_variante(pantalla.text)
    cargar(cliente, id_producto, {variantes[0]: 20})

    catalogo = cliente.get("/inventario").text
    assert "Polo Premium" in catalogo
    assert ">20<" in catalogo


def test_ce03_producto_simple_con_una_variante_base(cliente):
    crear_cuenta(cliente)
    pantalla = crear_producto(
        cliente, nombre="Gorra", atributo_1="", valores_1="", atributo_2="", valores_2=""
    )
    assert len(ids_de_variante(pantalla.text)) == 1
    assert "GOR-0001" in pantalla.text


def test_ce04_dos_atributos_de_siete_valores_dan_49_variantes(cliente):
    crear_cuenta(cliente)
    pantalla = crear_producto(cliente)
    variantes = ids_de_variante(pantalla.text)

    assert len(variantes) == 49
    assert len(set(variantes)) == 49  # sin duplicados
    # SKU unicos y correlativos.
    skus = set(re.findall(r"POL-\d{4}", pantalla.text))
    assert len(skus) == 49


def test_ce05_variantes_manuales_sin_generar_todas(cliente):
    crear_cuenta(cliente)
    pantalla = crear_producto(cliente, valores_1="M", valores_2="negro")
    assert len(ids_de_variante(pantalla.text)) == 1


# ---------------------------------------------------------------------------
# CE-06 a CE-10
# ---------------------------------------------------------------------------


def test_ce06_una_entrada_de_diez_aumenta_exactamente_diez(cliente):
    _, variantes = escenario_kova(cliente)
    cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "entrada", "cantidad": "10", "motivo": "Compra"},
        follow_redirects=True,
    )
    historial = cliente.get("/historial").text
    assert "20</td>" in historial  # stock anterior
    assert "30</td>" in historial  # stock posterior


def test_ce07_una_salida_de_dos_disminuye_exactamente_dos(cliente):
    _, variantes = escenario_kova(cliente)
    cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "salida", "cantidad": "2", "motivo": "Venta"},
        follow_redirects=True,
    )
    historial = cliente.get("/historial").text
    assert "-2" in historial
    assert "18</td>" in historial


def test_ce08_los_ajustes_aplican_el_delta_en_ambos_sentidos(cliente):
    _, variantes = escenario_kova(cliente)
    cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "ajuste_positivo", "cantidad": "3", "motivo": "Conteo fisico"},
        follow_redirects=True,
    )
    cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "ajuste_negativo", "cantidad": "5", "motivo": "Merma"},
        follow_redirects=True,
    )
    historial = cliente.get("/historial").text
    assert "Ajuste positivo" in historial
    assert "Ajuste negativo" in historial
    assert "18</td>" in historial  # 20 + 3 - 5


def test_ce09_stock_negativo_exige_confirmacion_y_motivo(cliente):
    _, variantes = escenario_kova(cliente)

    # Sin confirmar: no cambia nada.
    aviso = cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "salida", "cantidad": "50", "motivo": ""},
    )
    assert "deja el stock en -30" in aviso.text
    assert ">20<" in cliente.get("/inventario").text

    # Confirmando con motivo: se registra como incidencia.
    cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={
            "tipo": "salida",
            "cantidad": "50",
            "motivo": "Venta registrada tarde",
            "confirmar_negativo": "1",
        },
        follow_redirects=True,
    )
    incidencias = cliente.get("/historial?incidencias=1").text
    assert "Incidencia" in incidencias
    assert "Venta registrada tarde" in incidencias


def test_ce10_cancelar_compensa_restaura_y_no_se_repite(cliente):
    _, variantes = escenario_kova(cliente)
    cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "salida", "cantidad": "6", "motivo": "Venta"},
        follow_redirects=True,
    )
    historial = cliente.get("/historial").text
    id_movimiento = re.findall(r'/movimientos/([0-9a-f-]+)/cancelar', historial)[0]

    cliente.post(
        f"/movimientos/{id_movimiento}/cancelar",
        data={"motivo": "Se registro dos veces"},
        follow_redirects=True,
    )
    final = cliente.get("/historial").text
    assert "cancelado" in final
    assert "compensacion" in final
    assert ">20<" in cliente.get("/inventario").text  # stock restaurado

    # La segunda cancelacion se bloquea.
    segunda = cliente.post(
        f"/movimientos/{id_movimiento}/cancelar", data={"motivo": "Otra vez"}
    )
    assert segunda.status_code == 400
    assert "ya fue cancelado" in segunda.text


# ---------------------------------------------------------------------------
# CE-11 a CE-15
# ---------------------------------------------------------------------------


def test_ce11_los_datos_persisten_tras_cerrar_sesion(cliente):
    escenario_kova(cliente)
    cliente.post("/salir")

    cliente.post(
        "/entrar",
        data={"correo": "dueno@kova.com", "password": "clave_seguraKOVA"},
        follow_redirects=True,
    )
    catalogo = cliente.get("/inventario").text
    assert "Polo Premium" in catalogo
    assert ">33<" in catalogo  # 20 + 5 + 8


def test_ce12_las_metricas_coinciden_con_el_calculo_manual(cliente):
    escenario_kova(cliente)
    panel = cliente.get("/panel").text
    # 33 unidades x $250 = $8,250.
    assert ">33<" in panel
    assert "8,250.00" in panel


def test_ce13_solo_los_productos_bajo_su_minimo_aparecen_en_alertas(cliente):
    crear_cuenta(cliente)
    # Producto disponible: 20 unidades, minimo 5.
    pantalla = crear_producto(
        cliente,
        nombre="Gorra",
        minimo=5,
        atributo_1="",
        valores_1="",
        atributo_2="",
        valores_2="",
    )
    cargar(cliente, id_de_producto(pantalla.text), {ids_de_variante(pantalla.text)[0]: 20})

    # Producto agotado: sin existencias.
    crear_producto(cliente, nombre="Playera", valores_1="M", valores_2="negro")

    panel = cliente.get("/panel").text
    assert "Playera" in panel  # aparece en alertas
    seccion = panel.split("Requieren reposicion")[1].split("Mayor movimiento")[0]
    assert "Gorra" not in seccion  # la disponible no


def test_ce14_los_filtros_cambian_el_subconjunto(cliente):
    _, variantes = escenario_kova(cliente)
    cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "salida", "cantidad": "2", "motivo": "Venta"},
        follow_redirects=True,
    )

    def cuerpo_de_tabla(html: str) -> str:
        """Solo las filas de datos.

        Se compara contra el <tbody> y no contra la pagina completa porque los
        menus de filtro contienen las palabras 'Entrada' y 'Salida' como opciones,
        y buscarlas en todo el HTML daria un falso positivo.
        """
        return html.split("<tbody>")[-1].split("</tbody>")[0]

    def registros(html: str) -> int:
        return int(re.search(r"(\d+) registros", html).group(1))

    todos = cliente.get("/reportes?reporte=movimientos").text
    solo_salidas = cliente.get("/reportes?reporte=movimientos&tipo=salida").text

    assert "Entrada" in cuerpo_de_tabla(todos)
    assert "Entrada" not in cuerpo_de_tabla(solo_salidas)
    assert "Salida" in cuerpo_de_tabla(solo_salidas)
    assert registros(solo_salidas) < registros(todos)

    # Filtro por categoria inexistente devuelve vacio.
    ninguno = cliente.get("/reportes?reporte=inventario&categoria=NoExiste").text
    assert "Sin resultados" in ninguno
    assert registros(ninguno) == 0


def test_ce15_los_seis_reportes_abren_y_manejan_periodos_vacios(cliente):
    escenario_kova(cliente)
    for clave in ("inventario", "movimientos", "valor", "ranking", "reposicion", "flujo"):
        respuesta = cliente.get(f"/reportes?reporte={clave}")
        assert respuesta.status_code == 200, clave

    # Periodo sin movimientos: vacio con mensaje, sin error de division.
    vacio = cliente.get("/reportes?reporte=flujo&desde=2020-01-01&hasta=2020-01-31")
    assert vacio.status_code == 200
    assert "Sin resultados" in vacio.text


# ---------------------------------------------------------------------------
# CE-16 a CE-20
# ---------------------------------------------------------------------------


def test_ce16_el_excel_abre_y_refleja_el_filtro(cliente, tmp_path):
    escenario_kova(cliente)

    completo = cliente.get("/reportes/inventario.xlsx")
    assert completo.status_code == 200
    assert completo.headers["content-disposition"].endswith('.xlsx"')

    ruta = tmp_path / "inventario.xlsx"
    ruta.write_bytes(completo.content)
    hoja = load_workbook(ruta).active

    # Encabezados en la fila 5 y datos debajo.
    encabezados = [celda.value for celda in hoja[5]]
    assert "SKU" in encabezados
    assert "Valor" in encabezados
    # 49 variantes = 49 filas de datos.
    assert hoja.max_row == 5 + 49

    # El numero se guarda como numero, no como texto con simbolo de peso.
    columna_valor = encabezados.index("Valor") + 1
    celda = hoja.cell(row=6, column=columna_valor)
    assert isinstance(celda.value, (int, float))

    # Con filtro, el archivo trae menos filas.
    filtrado = cliente.get("/reportes/inventario.xlsx?categoria=NoExiste")
    ruta_filtrada = tmp_path / "filtrado.xlsx"
    ruta_filtrada.write_bytes(filtrado.content)
    assert load_workbook(ruta_filtrada).active.max_row < hoja.max_row


def test_ce17_el_pdf_se_descarga_legible(cliente):
    escenario_kova(cliente)
    respuesta = cliente.get("/reportes/inventario.pdf")

    assert respuesta.status_code == 200
    assert respuesta.headers["content-type"] == "application/pdf"
    assert respuesta.content.startswith(b"%PDF")
    assert len(respuesta.content) > 2000  # tiene contenido, no es un PDF vacio


def test_ce17b_pantalla_y_archivo_usan_el_mismo_subconjunto(cliente, tmp_path):
    """
    Riesgo de la planeacion: exportaciones distintas a los filtros.

    Se compara el numero de registros que dice la pantalla contra las filas del
    Excel generado con la misma query string.
    """
    escenario_kova(cliente)
    consulta = "reporte=inventario&estado=disponible"

    pantalla = cliente.get(f"/reportes?{consulta}").text
    registros = int(re.search(r"(\d+) registros", pantalla).group(1))
    assert registros > 0  # el filtro debe devolver algo, o la prueba no prueba nada

    ruta = tmp_path / "comparacion.xlsx"
    ruta.write_bytes(cliente.get(f"/reportes/inventario.xlsx?{consulta}").content)
    filas_excel = load_workbook(ruta).active.max_row - 5

    assert filas_excel == registros

    # Y con un filtro que no devuelve nada, ambos quedan vacios: el Excel trae
    # una sola fila, la del aviso, en lugar de datos.
    vacio = "reporte=inventario&estado=agotado"
    assert "Sin resultados" in cliente.get(f"/reportes?{vacio}").text

    ruta_vacia = tmp_path / "vacio.xlsx"
    ruta_vacia.write_bytes(cliente.get(f"/reportes/inventario.xlsx?{vacio}").content)
    hoja_vacia = load_workbook(ruta_vacia).active
    assert hoja_vacia.max_row == 6
    assert "Sin resultados" in str(hoja_vacia.cell(row=6, column=1).value)


def test_ce18_retirar_conserva_la_identidad_en_el_historial(cliente):
    id_producto, variantes = escenario_kova(cliente)
    cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "salida", "cantidad": "2", "motivo": "Venta"},
        follow_redirects=True,
    )
    cliente.post(f"/productos/{id_producto}/retirar", follow_redirects=True)

    assert "Polo Premium" not in cliente.get("/inventario").text
    # El historial conserva nombre y SKU.
    historial = cliente.get("/historial").text
    assert "Polo Premium" in historial
    assert "POL-0001" in historial
    # Y sigue consultable entre los retirados.
    assert "Polo Premium" in cliente.get("/inventario?retirados=1").text


def test_ce19_las_rutas_principales_responden(cliente):
    """
    Las seis rutas principales cargan sin error.

    La adaptacion visual a celular se verifica a mano; lo que si se puede
    automatizar es que ninguna ruta este rota y que el viewport este declarado.
    """
    escenario_kova(cliente)
    for ruta in ("/panel", "/inventario", "/historial", "/reportes", "/productos/nuevo"):
        respuesta = cliente.get(ruta)
        assert respuesta.status_code == 200, ruta
        assert 'name="viewport"' in respuesta.text, ruta


def test_ce20_no_hay_rutas_rotas_ni_operaciones_parciales(cliente):
    """
    Toda ruta protegida redirige si no hay sesion, y ninguna deja datos a medias.

    Se comprueba que un intento invalido no altere el inventario.
    """
    # Sin sesion, nada de lo protegido responde 200.
    for ruta in ("/panel", "/inventario", "/historial", "/reportes"):
        assert cliente.get(ruta, follow_redirects=False).status_code == 303, ruta

    _, variantes = escenario_kova(cliente)
    antes = cliente.get("/inventario").text

    # Cantidad invalida: error explicado, inventario intacto.
    fallo = cliente.post(
        f"/variantes/{variantes[0]}/movimiento",
        data={"tipo": "salida", "cantidad": "abc", "motivo": ""},
    )
    assert fallo.status_code == 400
    assert cliente.get("/inventario").text == antes
