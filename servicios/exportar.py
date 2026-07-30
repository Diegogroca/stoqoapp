"""
Exportacion a Excel y PDF (Etapa 6).

Ambas funciones reciben el MISMO objeto `Reporte` que se dibuja en pantalla. No
vuelven a consultar la base de datos ni reciben filtros: si lo hicieran, podrian
producir un subconjunto distinto al que el usuario vio, que es exactamente el
riesgo que la planeacion marcaba como alto.

Detalle de Excel que importa: los numeros se escriben como numeros, no como texto
con simbolo de peso incrustado. Una columna con "$1,250.00" es inservible para
sumar o graficar; el formato se aplica al estilo de la celda, no al valor.
"""

from __future__ import annotations

import io
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from servicios.reportes import Filtros, Reporte

ZONA = ZoneInfo("America/Mexico_City")

# Identidad visual de Stoqo, la misma de la interfaz.
TINTA = "0B1B2B"
ACENTO = "0E9384"
BORDE = "DFE6E9"

COLUMNAS_DINERO = {"costo", "valor"}


def _sello_de_tiempo() -> str:
    return datetime.now(ZONA).strftime("%d/%m/%Y %H:%M")


def a_excel(reporte: Reporte, filtros: Filtros, empresa: str) -> bytes:
    """Genera un libro de Excel en memoria y devuelve sus bytes."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = reporte.titulo[:31] or "Reporte"  # Excel limita a 31 caracteres

    hoja["A1"] = f"{empresa} - {reporte.titulo}"
    hoja["A1"].font = Font(bold=True, size=14, color=TINTA)
    hoja["A2"] = f"Filtros: {filtros.descripcion()}"
    hoja["A3"] = f"Generado: {_sello_de_tiempo()}"
    for celda in ("A2", "A3"):
        hoja[celda].font = Font(size=9, color="7C8B99")

    fila_encabezado = 5
    for indice, columna in enumerate(reporte.columnas, start=1):
        celda = hoja.cell(row=fila_encabezado, column=indice, value=columna)
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = PatternFill("solid", fgColor=ACENTO)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    indices_dinero = {
        indice
        for indice, nombre in enumerate(reporte.columnas)
        if nombre.strip().lower() in COLUMNAS_DINERO
    }

    for desplazamiento, fila in enumerate(reporte.filas, start=1):
        for indice, valor in enumerate(fila):
            celda = hoja.cell(
                row=fila_encabezado + desplazamiento, column=indice + 1, value=valor
            )
            if indice in indices_dinero and isinstance(valor, (int, float)):
                # El formato es del estilo; el valor sigue siendo numerico.
                celda.number_format = '"$"#,##0.00'
            if indice in reporte.alineadas_derecha:
                celda.alignment = Alignment(horizontal="right")

    # Ancho segun el contenido mas largo de cada columna, con tope razonable.
    for indice, columna in enumerate(reporte.columnas, start=1):
        largo = max(
            [len(str(columna))]
            + [len(str(fila[indice - 1])) for fila in reporte.filas]
            or [10]
        )
        hoja.column_dimensions[get_column_letter(indice)].width = min(largo + 4, 42)

    # Fila de encabezado congelada: al bajar en un reporte largo se siguen viendo
    # los nombres de las columnas.
    hoja.freeze_panes = hoja.cell(row=fila_encabezado + 1, column=1)

    if reporte.vacio:
        hoja.cell(
            row=fila_encabezado + 1,
            column=1,
            value="Sin resultados para los filtros aplicados.",
        )

    memoria = io.BytesIO()
    libro.save(memoria)
    return memoria.getvalue()


def a_pdf(reporte: Reporte, filtros: Filtros, empresa: str) -> bytes:
    """
    Genera un PDF en memoria y devuelve sus bytes.

    Horizontal y no vertical: los reportes de inventario tienen muchas columnas y
    en vertical se cortarian.
    """
    memoria = io.BytesIO()
    documento = SimpleDocTemplate(
        memoria,
        pagesize=landscape(letter),
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
        title=f"{empresa} - {reporte.titulo}",
        author="Stoqo",
    )

    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        "TituloStoqo",
        parent=estilos["Title"],
        fontSize=16,
        alignment=0,
        textColor=colors.HexColor(f"#{TINTA}"),
        spaceAfter=4,
    )
    menor = ParagraphStyle(
        "MenorStoqo",
        parent=estilos["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#7C8B99"),
    )
    celda = ParagraphStyle("CeldaStoqo", parent=estilos["Normal"], fontSize=7.5, leading=9.5)

    contenido = [
        Paragraph(f"{empresa} &middot; {reporte.titulo}", titulo),
        Paragraph(f"Filtros: {filtros.descripcion()}", menor),
        Paragraph(f"Generado: {_sello_de_tiempo()}", menor),
        Spacer(1, 0.5 * cm),
    ]

    if reporte.vacio:
        contenido.append(
            Paragraph("Sin resultados para los filtros aplicados.", estilos["Normal"])
        )
    else:
        datos = [[Paragraph(f"<b>{c}</b>", celda) for c in reporte.columnas]]
        for fila in reporte.filas:
            datos.append(
                [
                    Paragraph(
                        f"{valor:,.2f}" if isinstance(valor, float) else str(valor),
                        celda,
                    )
                    for valor in fila
                ]
            )

        tabla = Table(datos, repeatRows=1)
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{ACENTO}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(f"#{BORDE}")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7F8")]),
        ]
        for indice in reporte.alineadas_derecha:
            estilo.append(("ALIGN", (indice, 1), (indice, -1), "RIGHT"))
        tabla.setStyle(TableStyle(estilo))
        contenido.append(tabla)

    if reporte.nota:
        contenido.append(Spacer(1, 0.4 * cm))
        contenido.append(Paragraph(reporte.nota, menor))

    contenido.append(Spacer(1, 0.3 * cm))
    contenido.append(
        Paragraph(
            f"{len(reporte.filas)} registros &middot; Stoqo, proyecto final de "
            "Python para negocios",
            menor,
        )
    )

    documento.build(contenido)
    return memoria.getvalue()


def nombre_archivo(reporte: Reporte, extension: str) -> str:
    """Nombre descargable con fecha: stoqo-inventario-2026-07-29.xlsx"""
    fecha = datetime.now(ZONA).strftime("%Y-%m-%d")
    return f"stoqo-{reporte.clave}-{fecha}.{extension}"
